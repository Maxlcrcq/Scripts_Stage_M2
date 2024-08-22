import os

import openpyxl
import pandas as pd

# Chemin du dossier contenant les fichiers .txt
txt_dossier = r"C:\Users\lecle\Desktop\Clean\9 - AMR\Txt"

# Chemin du dossier contenant les fichiers .xlsx
xlsx_dossier = r"C:\Users\lecle\Desktop\Clean\9 - AMR\Xlsx"

# Chemin du dossier où seront stockés les fichiers filtrés
filtred_dossier = r"C:\Users\lecle\Desktop\Clean\9 - AMR\Filtred"

# Renommer les fichiers .txt en remplaçant "_out" par ""
for filename in os.listdir(txt_dossier):
    if filename.endswith(".txt"):
        new_filename = filename.replace("_out", "")
        os.rename(os.path.join(txt_dossier, filename), os.path.join(txt_dossier, new_filename))

# Parcourir tous les fichiers du dossier
for fichier in os.listdir(txt_dossier):
    if fichier.endswith(".txt"):
        # Chemin complet du fichier .txt
        chemin_fichier = os.path.join(txt_dossier, fichier)
        
        # Lire le fichier .txt en utilisant le caractère de tabulation comme séparateur
        df = pd.read_csv(chemin_fichier, sep="\t")
        
        # Chemin du fichier Excel de sortie
        chemin_excel = os.path.join(xlsx_dossier, os.path.splitext(fichier)[0] + ".xlsx")
        
        # Enregistrer le DataFrame dans un fichier Excel
        df.to_excel(chemin_excel, index=False)

# Parcourir tous les fichiers du dossier
for fichier in os.listdir(xlsx_dossier):
    if fichier.endswith(".xlsx"):
        file_path = os.path.join(xlsx_dossier, fichier)
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        # Extraire le nom de fichier sans extension
        file_name = os.path.splitext(fichier)[0]
        
        # Insérer une colonne "Souche" à la première position
        sheet.insert_cols(1)
        sheet.cell(row=1, column=1, value="Souche")
        
        # Vérifier si le fichier est vide (à l'exception de l'en-tête)
        max_row = sheet.max_row
        if max_row == 1:
            # La feuille est vide, écrire le nom du fichier (souche) à la première ligne
            sheet.cell(row=2, column=1, value=file_name)
        
        else:
            # Ecrire le nom du fichier (souche) dans toutes les lignes
            for row_num in range(2, max_row + 1):  # Démarre à 2 pour éviter l'en-tête
                sheet.cell(row=row_num, column=1, value=file_name)
        
        wb.save(file_path)
        wb.close()

# Colonnes à récupérer dans les fichiers d'entrée
columns_to_retrieve = ['Souche','Class', 'Subclass', 'Sequence name', 'Gene symbol']

# Parcourir tous les fichiers du dossier
for fichier in os.listdir(xlsx_dossier):
    if fichier.endswith('.xlsx'):
        # Construire le nom de fichier de sortie
        nom_fichier, extension = os.path.splitext(fichier)
        
        if nom_fichier.startswith('GCF_'):
            # Pour les fichiers commençant par 'GCF_', on ne coupe pas après le premier '_'
            nouveau_fichier = f"{nom_fichier}_AMR.xlsx"  # Par exemple, GCF_002814135_AMR.xlsx
        elif nom_fichier.startswith('II_'):
            # Pour les fichiers commençant par 'II_', on ne coupe pas après le premier '_'
            nouveau_fichier = f"{nom_fichier}_AMR.xlsx"
        else:
            # Pour les autres fichiers, on coupe après le premier '_'
            nouveau_fichier = f"{nom_fichier.split('_')[0]}_AMR.xlsx"  # Par exemple, A1_AMR.xlsx si l'entrée est A1_out.xlsx

        # Lire le fichier Excel
        df = pd.read_excel(os.path.join(xlsx_dossier, fichier))

        # Sélectionner uniquement les colonnes d'intérêt
        selected_columns = df[columns_to_retrieve]

        # Trier le DataFrame par certaines colonnes (par exemple, 'Souche')
        sorted_columns = selected_columns.sort_values(by=['Souche'])

        # Enregistrer le nouveau fichier avec le nom construit dans le dossier Filtred
        sorted_columns.to_excel(os.path.join(filtred_dossier, nouveau_fichier), index=False)

# Merge tous les fichiers Excel en un seul DataFrame
dossier = filtred_dossier
# Liste pour stocker les données de chaque fichier Excel
donnees = []

# Parcourir tous les fichiers du dossier
for fichier in os.listdir(dossier):
    if fichier.endswith("_AMR.xlsx"):
        # Chemin complet du fichier
        chemin_fichier = os.path.join(dossier, fichier)
        
        # Lire le fichier Excel et ajouter les données à la liste
        df = pd.read_excel(chemin_fichier)
        donnees.append(df)

# Fusionner toutes les données en un seul DataFrame
donnees_fusionnees = pd.concat(donnees)

# Enregistrer les données fusionnées dans un nouveau fichier Excel
donnees_fusionnees.to_excel(os.path.join(dossier, "AMR_fusionne.xlsx"), index=False)

# Lire le fichier Excel AMR_fusionne
df = pd.read_excel(os.path.join(filtred_dossier, "AMR_fusionne.xlsx"))

# Filtrer les lignes qui ont une valeur pour la colonne "Souche" mais pas pour les autres colonnes
df_filtered = df[(df["Class"].notna()) | (df["Subclass"].notna()) | (df["Sequence name"].notna()) | (df["Gene symbol"].notna())]

# Enregistrer les données filtrées dans un nouveau fichier Excel
df_filtered.to_excel(os.path.join(filtred_dossier, "AMR_filtre.xlsx"), index=False)